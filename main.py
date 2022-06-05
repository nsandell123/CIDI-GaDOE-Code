import pandas as pd
import pandas.io.formats.excel
pandas.io.formats.excel.ExcelFormatter.header_style = None
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
import os
import os.path
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import glob
from selenium.webdriver.common.by import By
import win32com.client as win32
os.environ['PATH'] += r";C:\SeleniumDriver"
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = True
excel.DisplayAlerts = False

### Initial Scraping

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://gatfl.gatech.edu/sri/users/login")
driver.maximize_window()


username = driver.find_element(By.ID, "UserUsername")
username.clear()
username.send_keys("nsandell6")


password = driver.find_element(By.ID, "UserPassword")
password.clear()
password.send_keys("nsandell123!")


button = driver.find_element(By.XPATH, "//input[@type='submit']")
button.click()

requests_link = driver.find_element(By.LINK_TEXT, "GA-DOE Requests")
requests_link.click()

export_requests = driver.find_element(By.LINK_TEXT, 'Export all')
export_requests.click()

driver.get('https://gatfl.gatech.edu/sri/users')

export_requests = driver.find_element(By.LINK_TEXT, 'Export GA-DoE')
export_requests.click()


### Collect CSV from downloads
name = str(input("What is your account name on Windows?"))
#r'C:\Users\sande\Downloads'
file_system_base = "C:\\Users\\{name}".format(name=name)
downloads_folder = r"\Downloads"
downloads_folder_path = file_system_base + downloads_folder
file_type = '\*csv'
files = glob.glob(downloads_folder_path + file_type)

files = sorted(files, key = os.path.getctime, reverse=True)
# files[0] is users and files[1] is requests

### Insert into Excel
df_requests = pd.read_csv(files[1])
df_users = pd.read_csv(files[0])

os.remove(files[0])
os.remove(files[1])

path = str(input("What is the location of the CDR?"))
wb = load_workbook(filename=path, read_only=False, keep_vba=True)

date = str(input("What is today's date in (MMDDYY)"))
for sheet in wb.sheetnames:
    if 'Combined Data' in sheet:
        wb[sheet].title = 'Combined Data ' + date
ws_requests = wb.create_sheet(index=2)
ws_requests.title = "GA DoE Requests " + date

wb.remove(wb[wb.sheetnames[4]])

ws_users = wb.create_sheet(index=4)
ws_users.title = "GA DoE Users " + date

for row in dataframe_to_rows(df_requests, index=True, header=True):
    ws_requests.append(row)

ws_requests.delete_rows(2)
ws_requests.delete_cols(1)

for row in dataframe_to_rows(df_users, index=True, header=True):
    ws_users.append(row)
ws_users.delete_rows(2)
ws_users.delete_cols(1)

wb.remove(wb[wb.sheetnames[6]])

new_file_name = os.getcwd() + "\\" + "Combined Data " + date + ".xlsm"
wb.save(filename=new_file_name)

# macro_wb = excel.Workbooks.Open(new_file_name)
# excel.Run("\'" + new_file_name + "\'" + "!Module1.GlobalChecking")
# excel.Run("\'" + new_file_name + "\'" + "!Module2.RequestsFormatting")
# excel.Run("\'" + new_file_name + "\'" + "!Module2.UsersFormatting")
# excel.Run("\'" + new_file_name + "\'" + "!Module3.UsersDataCleaning")
# excel.Run("\'" + new_file_name + "\'" + "!Module4.FindNewUsers")
# excel.Run("\'" + new_file_name + "\'" + "!Module5.FindRequestsDiff")
# excel.Run("\'" + new_file_name + "\'" + "!Module6.UpdateTable")
# excel.Run("\'" + new_file_name + "\'" + "!Module7.UpdateRegisterSheets")
